#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据处理模块，负责读取和处理Excel文件中的班级名册数据
使用 openpyxl 直接读取，无 pandas 依赖
"""

from typing import Dict, List, Optional, Any
from collections import Counter
from openpyxl import load_workbook


class DataHandler:
    """数据处理类"""

    def __init__(self, file_path):
        self.file_path = file_path
        self._data: List[Dict[str, Any]] = []
        self.columns: List[str] = []
        self.column_metadata: Dict[str, Dict] = {}
        self.load_data()

    def load_data(self):
        """读取Excel文件数据"""
        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if not rows:
                raise ValueError("数据源文件为空")

            self.columns = [str(c) if c is not None else f"列{i}" for i, c in enumerate(rows[0])]
            self._data = []

            for row in rows[1:]:
                if all(cell is None for cell in row):
                    continue
                self._data.append(dict(zip(self.columns, row)))

            if not self._data:
                raise ValueError("数据源文件为空")

            self._generate_column_metadata()
            print(f"成功读取数据，共{len(self._data)}条记录，包含{len(self.columns)}列")
        except FileNotFoundError:
            raise FileNotFoundError(f"数据源文件不存在: {self.file_path}")
        except PermissionError:
            raise PermissionError(f"无权限访问数据源文件: {self.file_path}")
        except Exception as e:
            if "数据源文件为空" in str(e):
                raise
            raise ValueError(f"读取数据失败: {str(e)}. 请检查文件格式和内容是否正确")

    def reload_data(self):
        """重新加载数据，用于刷新数据源"""
        print("正在刷新数据...")
        self.load_data()

    def _generate_column_metadata(self):
        """生成列元数据，包括列名、类型和唯一值"""
        self.column_metadata = {}
        for col in self.columns:
            values = [row[col] for row in self._data if row[col] is not None]
            unique_values = list(dict.fromkeys(values))
            if len(unique_values) > 20:
                unique_values = unique_values[:20] + ["..."]

            self.column_metadata[col] = {
                'name': col,
                'type': self._detect_type(values),
                'unique_values': unique_values,
                'non_null_count': len(values),
                'total_count': len(self._data)
            }

    @staticmethod
    def _detect_type(values: list) -> str:
        if not values:
            return "object"
        sample = values[0]
        if isinstance(sample, (int, float)):
            return "number"
        return "object"

    def get_all_data(self) -> List[Dict[str, Any]]:
        """获取所有数据（返回副本）"""
        return [row.copy() for row in self._data]

    def get_columns(self) -> List[str]:
        """获取列名列表"""
        return self.columns.copy()

    def get_column_metadata(self) -> Dict[str, Dict]:
        """获取列元数据"""
        return {k: v.copy() for k, v in self.column_metadata.items()}

    def filter_data(self, filters: Optional[Dict[str, Any]] = None, **kwargs) -> List[Dict[str, Any]]:
        """
        根据条件筛选数据

        Args:
            filters: 筛选条件字典，键为列名，值为筛选值

        Returns:
            筛选后的数据列表
        """
        result = list(self._data)

        if filters:
            for col, value in filters.items():
                if col in self.columns and value is not None:
                    col_values = set(row[col] for row in self._data if row[col] is not None)
                    if value in col_values:
                        result = [row for row in result if row.get(col) == value]
                    else:
                        return []

        return result

    def validate_filters(self, filters: Dict[str, Any], strict: bool = False) -> List[str]:
        """
        验证筛选条件的有效性

        Args:
            filters: 筛选条件字典
            strict: 是否严格验证

        Returns:
            错误信息列表，为空表示验证通过
        """
        errors = []
        for col, value in filters.items():
            if col not in self.columns:
                errors.append(f"列 '{col}' 不存在于当前数据源中")
            elif strict:
                col_values = set(row[col] for row in self._data if row[col] is not None)
                if value not in col_values:
                    errors.append(f"值 '{value}' 不在列 '{col}' 的有效值列表中")
        return errors

    def get_statistics(self) -> Dict[str, Any]:
        """获取数据统计信息"""
        stats: Dict[str, Any] = {'总人数': len(self._data)}

        for col in self.columns:
            values = [row[col] for row in self._data]
            non_null = [v for v in values if v is not None]
            stats[f'{col}_非空数量'] = len(non_null)

            counter = Counter(non_null)
            if len(counter) <= 10:
                for value, count in counter.items():
                    stats[f'{col}_{value}'] = count

        return stats

    @property
    def df(self):
        """兼容旧接口，返回自身作为 DataFrame-like 对象"""
        return _DataFrameProxy(self._data, self.columns)


class _DataFrameProxy:
    """轻量级 DataFrame 兼容层，提供 iterrows / columns / empty / head / equals 等接口"""

    def __init__(self, data: List[Dict[str, Any]], columns: List[str]):
        self._data = data
        self.columns = columns

    @property
    def empty(self) -> bool:
        return len(self._data) == 0

    def __len__(self) -> int:
        return len(self._data)

    def head(self, n: int) -> '_DataFrameProxy':
        return _DataFrameProxy(self._data[:n], self.columns)

    def copy(self) -> '_DataFrameProxy':
        return _DataFrameProxy([row.copy() for row in self._data], self.columns)

    def iterrows(self):
        for i, row in enumerate(self._data):
            yield i, _SeriesProxy(row)

    def __iter__(self):
        return iter(self._data)

    def __getitem__(self, key):
        if isinstance(key, str):
            return [row.get(key) for row in self._data]
        if isinstance(key, list):
            return _DataFrameProxy(
                [{k: row[k] for k in key if k in row} for row in self._data],
                key
            )
        raise TypeError(f"Unsupported key type: {type(key)}")

    def equals(self, other: '_DataFrameProxy') -> bool:
        if len(self._data) != len(other._data):
            return False
        return all(a == b for a, b in zip(self._data, other._data))

    def drop_duplicates(self, subset: Optional[List[str]] = None, keep: str = 'first') -> '_DataFrameProxy':
        seen = set()
        result = []
        items = self._data if keep == 'first' else reversed(self._data)
        for row in items:
            if subset:
                key = tuple(row.get(c) for c in subset)
            else:
                key = tuple(row.values())
            if key not in seen:
                seen.add(key)
                result.append(row)
        if keep == 'last':
            result.reverse()
        return _DataFrameProxy(result, self.columns)

    def isin(self, values) -> List[bool]:
        return [row in values for row in self._data]

    def __bool__(self) -> bool:
        return len(self._data) > 0


class _SeriesProxy:
    """轻量级 Series 兼容层，提供 dict-like 接口"""

    def __init__(self, data: Dict[str, Any]):
        self._data = data

    def get(self, key, default=None):
        return self._data.get(key, default)

    def __getitem__(self, key):
        return self._data[key]

    def __contains__(self, key):
        return key in self._data

    @property
    def index(self):
        return self._data.keys()
